#!/usr/bin/env python3
"""
09_enrichment.py
================
Step 9 of the TT4 FAPESP case-control snRNA-seq pipeline (MTLE-HS vs autopsy).

STRATEGIC (not exhaustive) enrichment, per plan agreed with the PI:

  1. Genes MT-* (mitochondrial) REMOVED before enrichment -- they dominate
     nearly every DE list due to PMI confound (autopsy has PMI, caso doesn't)
     and would flood every comparison with "oxidative phosphorylation" noise.
  2. Uses LEVEL 2 (robust: FDR<0.05, |logFC|>0.5, protein_coding) gene lists,
     at cell_type level (Moment 1) -- not Level 1, not Leiden subclusters.
  3. Perspectiva I (caso vs controle, todos) -- FULL enrichment, all priority
     cell types.
  4. Perspectiva I-G1 / I-G2 -- RESTRICTED to cell types that were ALSO
     significant (Level 2) in Perspectiva I (confirmatory, not exploratory).
  5. Perspectiva II (G1 vs G2, caso only) -- NO formal enrichment (0 genes
     Level 2 in every cell type). Only the Level 1 gene list is exported as
     a hypothesis-generating table.
  6. Priority cell types only: Astrocyte, Microglia, Oligodendrocyte,
     Inhibitory Neuron, Excitatory Neuron (T Cell / Ependymal / Endothelial
     excluded -- too few genes for meaningful enrichment).
  7. Up- and down-regulated genes enriched SEPARATELY.

Output
------
    tables/enrichment/<tag>/<cell_type>_<direction>_{GO_BP,Reactome}.csv
    tables/enrichment/<tag>/enrichment_summary.csv
    figures/enrichment/<tag>/dotplot_<cell_type>_<direction>.png / .pdf
    figures/enrichment/enrichment_strategic.xlsx   -- one workbook, all comparisons
    tables/de_cataloged/G1_vs_G2_caso/nivel1_genes_hypothesis.csv  (no enrichment)

Run
---
    conda activate scrna
    screen -S step09
    nohup python 09_enrichment.py > logs/09_enrichment.log 2>&1 &
    tail -f logs/09_enrichment.log
"""

import os
import sys
import re
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import gseapy as gp

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.environ["CUDA_VISIBLE_DEVICES"] = ""

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import TABLE_DIR, FIG_DIR

CATALOG_DIR = TABLE_DIR / "de_cataloged"
ENRICH_DIR  = TABLE_DIR / "enrichment"
FIG_ENR_DIR = FIG_DIR / "enrichment"
ENRICH_DIR.mkdir(parents=True, exist_ok=True)
FIG_ENR_DIR.mkdir(parents=True, exist_ok=True)

GENE_SETS = {
    "GO_BP":    "GO_Biological_Process_2023",
    "Reactome": "Reactome_2022",
}
FDR_CUT = 0.05

PRIORITY_CELL_TYPES = [
    "Astrocyte", "Microglia", "Oligodendrocyte",
    "Inhibitory Neuron", "Excitatory Neuron",
]

MT_PATTERN = re.compile(r"^MT-", re.IGNORECASE)

def strip_mt_genes(genes):
    return [g for g in genes if not MT_PATTERN.match(str(g))]

# =============================================================================
# COMPARISONS
# =============================================================================

COMPARISONS = [
    dict(tag="caso_vs_controle_all", title="Perspectiva I", mode="full"),
    dict(tag="caso_vs_controle_G1",  title="Perspectiva I-G1", mode="restricted"),
    dict(tag="caso_vs_controle_G2",  title="Perspectiva I-G2", mode="restricted"),
    dict(tag="G1_vs_G2_caso",        title="Perspectiva II", mode="hypothesis_only"),
]

print("=" * 60)
print("09_enrichment.py — estrategico")
print("=" * 60)

# =============================================================================
# HELPERS
# =============================================================================

def load_level2_by_celltype(tag):
    """Load Level 2 genes for Moment 1 (cell_type), split by cluster/direction,
    with MT-* genes removed."""
    path = CATALOG_DIR / tag / "nivel2_robusto_todos.csv"
    if not path.exists():
        return {}
    df = pd.read_csv(path, low_memory=False)
    df = df[df["momento"] == 1]  # Moment 1 = cell_type
    label_col = "symbol" if "symbol" in df.columns else "gene"
    out = {}
    for ct in df["cluster"].unique():
        if ct not in PRIORITY_CELL_TYPES:
            continue
        sub = df[df["cluster"] == ct]
        for direction in ["up", "down"]:
            genes = sub[sub["direction"] == direction][label_col].dropna().unique().tolist()
            genes = strip_mt_genes(genes)
            out[(ct, direction)] = sorted(set(genes))
    return out


def load_background(tag):
    path = CATALOG_DIR / tag / "de_all_annotated.csv"
    if not path.exists():
        return []
    df = pd.read_csv(path, low_memory=False)
    df = df[df["momento"] == 1]
    label_col = "symbol" if "symbol" in df.columns else "gene"
    bg = df[df["biotype_clean"].notna()][label_col].dropna().unique().tolist()
    return strip_mt_genes(bg)


def run_enrichment(gene_list, background, direction, ct, tag):
    """Run Enrichr for GO_BP + Reactome, FDR<0.05. Returns dict {gs_label: df}."""
    results = {}
    for gs_label, gs_name in GENE_SETS.items():
        if len(gene_list) < 5:
            results[gs_label] = pd.DataFrame()
            continue
        try:
            enr = gp.enrichr(
                gene_list=gene_list, gene_sets=gs_name, background=background,
                organism="human", outdir=None, cutoff=FDR_CUT, verbose=False,
            )
            df = enr.results.copy().rename(columns={
                "Term": "term", "P-value": "pvalue", "Adjusted P-value": "fdr",
                "Odds Ratio": "odds_ratio", "Combined Score": "combined_score",
                "Genes": "genes",
            })
            df["direction"]   = direction
            df["cell_type"]   = ct
            df["comparison"]  = tag
            df["gene_set"]    = gs_label
            df["n_input"]     = len(gene_list)
            df["n_overlap"]   = df["genes"].apply(
                lambda x: len(str(x).split(";")) if pd.notna(x) and x != "" else 0
            )
            df_sig = df[df["fdr"] < FDR_CUT].sort_values("fdr")
            results[gs_label] = df_sig
        except Exception as e:
            print(f"      ERROR {gs_label} ({ct}/{direction}): {e}")
            results[gs_label] = pd.DataFrame()
    return results


def dotplot(df, title, outpath, top_n=15):
    if df is None or len(df) == 0:
        return
    df = df.head(top_n).copy()
    df["term_short"] = df["term"].str.replace(r"\s*\(GO:\d+\)", "", regex=True).str[:60]
    df["-log10_fdr"] = -np.log10(df["fdr"].clip(lower=1e-10))

    fig, ax = plt.subplots(figsize=(8, max(3.5, len(df) * 0.35)))
    scatter = ax.scatter(
        x=df["-log10_fdr"], y=range(len(df)),
        s=df["n_overlap"] * 12, c=df["odds_ratio"],
        cmap="RdBu_r", alpha=0.85, edgecolors="grey", linewidths=0.4,
    )
    ax.set_yticks(range(len(df)))
    ax.set_yticklabels(df["term_short"], fontsize=8)
    ax.invert_yaxis()
    ax.set_xlabel("-log10(FDR)", fontsize=10)
    ax.set_title(title, fontsize=10, pad=10)
    ax.axvline(x=-np.log10(0.05), color="gray", linestyle="--", linewidth=0.8, alpha=0.6)
    cbar = plt.colorbar(scatter, ax=ax, pad=0.02)
    cbar.set_label("Odds Ratio", fontsize=9)
    plt.tight_layout()
    plt.savefig(f"{outpath}.png", dpi=200, bbox_inches="tight")
    plt.savefig(f"{outpath}.pdf", bbox_inches="tight")
    plt.close()


# =============================================================================
# EXCEL HELPERS (simplified single workbook, all comparisons)
# =============================================================================

FONT_NAME = "Calibri"
HDR_UP, FILL_UP     = "C0392B", "FADBD8"
HDR_DOWN, FILL_DOWN = "2471A3", "D6EAF8"
FILL_ALT = "F2F2F2"
THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DISPLAY_COLS = ["term", "n_overlap", "n_input", "odds_ratio", "fdr", "genes"]
COL_LABELS = {"term": "Term", "n_overlap": "N genes", "n_input": "N input",
              "odds_ratio": "Odds Ratio", "fdr": "FDR", "genes": "Genes"}

def style_header(c, hdr_color):
    c.font = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=hdr_color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

def style_data(c, row_fill=None):
    c.font = Font(name=FONT_NAME, size=9)
    c.border = BORDER
    c.alignment = Alignment(vertical="center")
    if row_fill:
        c.fill = PatternFill("solid", start_color=row_fill)

def write_sheet(wb, sheet_name, sections):
    """sections: list of (label, df, hdr_color, fill_color)"""
    ws = wb.create_sheet(title=sheet_name[:31])
    row = 1
    for label, df, hdr_color, fill_color in sections:
        ws.merge_cells(f"A{row}:{get_column_letter(len(DISPLAY_COLS))}{row}")
        c = ws.cell(row=row, column=1)
        c.value = label
        c.font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=hdr_color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        row += 1
        if df is None or len(df) == 0:
            ws.cell(row=row, column=1).value = "Sem termos significativos (FDR<0.05)"
            ws.cell(row=row, column=1).font = Font(name=FONT_NAME, italic=True, color="888888", size=9)
            row += 2
            continue
        for col_idx, col in enumerate(DISPLAY_COLS, start=1):
            c = ws.cell(row=row, column=col_idx)
            c.value = COL_LABELS.get(col, col)
            style_header(c, hdr_color)
        row += 1
        cols_present = [c for c in DISPLAY_COLS if c in df.columns]
        for i, (_, r) in enumerate(df[cols_present].head(15).iterrows()):
            zebra = FILL_ALT if i % 2 else None
            for col_idx, col in enumerate(cols_present, start=1):
                c = ws.cell(row=row, column=col_idx)
                c.value = r[col] if pd.notna(r[col]) else ""
                style_data(c, row_fill=zebra)
            row += 1
        for col_idx, col in enumerate(cols_present, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 55 if col == "term" else (40 if col == "genes" else 13)
        row += 1
    return ws


# =============================================================================
# MAIN LOOP
# =============================================================================

wb = Workbook()
wb.remove(wb.active)

# Track which cell types were significant (Level2) in Perspectiva I, to restrict
# I-G1 / I-G2 to the confirmatory subset.
sig_celltypes_perspective1 = set()

all_summary_rows = []

for cfg in COMPARISONS:
    tag, title, mode = cfg["tag"], cfg["title"], cfg["mode"]
    print(f"\n{'='*60}\n{title}  ({tag})  [mode={mode}]\n{'='*60}")

    tag_table_dir = ENRICH_DIR / tag
    tag_fig_dir   = FIG_ENR_DIR / tag
    tag_table_dir.mkdir(parents=True, exist_ok=True)
    tag_fig_dir.mkdir(parents=True, exist_ok=True)

    if mode == "hypothesis_only":
        # Perspectiva II: no formal enrichment, export Level1 gene list only
        l1_path = CATALOG_DIR / tag / "nivel1_exploratorio_todos.csv"
        if l1_path.exists():
            l1 = pd.read_csv(l1_path, low_memory=False)
            l1 = l1[l1["momento"] == 1]
            label_col = "symbol" if "symbol" in l1.columns else "gene"
            l1["is_mt"] = l1[label_col].astype(str).str.match(MT_PATTERN)
            l1_clean = l1[~l1["is_mt"]]
            out_path = CATALOG_DIR / tag / "nivel1_genes_hypothesis.csv"
            l1_clean[[label_col, "cluster", "direction", "logFC", "pvalue"]].to_csv(
                out_path, index=False
            )
            print(f"  Sem enrichment formal (0 genes Level 2 em todos os tipos).")
            print(f"  Lista Level 1 (hipotese-geradora, MT-* removidos) salva: {out_path}")
            print(f"  N genes: {len(l1_clean)} (across cell types)")
        else:
            print(f"  AVISO: {l1_path} nao encontrado.")
        continue

    gene_lists = load_level2_by_celltype(tag)
    background = load_background(tag)
    print(f"  Background: {len(background):,} genes (MT-* removidos)")

    cell_types_here = sorted(set(ct for ct, _ in gene_lists.keys()))

    if mode == "restricted":
        cell_types_here = [ct for ct in cell_types_here if ct in sig_celltypes_perspective1]
        print(f"  Modo confirmatorio: restrito aos tipos significativos na Perspectiva I: "
              f"{cell_types_here or '(nenhum)'}")

    sections_by_celltype = {}

    for ct in PRIORITY_CELL_TYPES:
        if ct not in cell_types_here:
            continue
        for direction in ["up", "down"]:
            genes = gene_lists.get((ct, direction), [])
            if len(genes) < 5:
                print(f"  [SKIP] {ct} / {direction}: apenas {len(genes)} genes (min 5)")
                continue
            print(f"  Enrichment: {ct} / {direction} ({len(genes)} genes)...")
            res = run_enrichment(genes, background, direction, ct, tag)

            for gs_label, df in res.items():
                out_csv = tag_table_dir / f"{ct.replace(' ','_')}_{direction}_{gs_label}.csv"
                df.to_csv(out_csv, index=False)
                n_sig = len(df)
                print(f"    {gs_label}: {n_sig} termos significativos (FDR<{FDR_CUT})")
                all_summary_rows.append({
                    "comparison": tag, "cell_type": ct, "direction": direction,
                    "gene_set": gs_label, "n_input_genes": len(genes), "n_sig_terms": n_sig,
                })
                if mode == "full" and n_sig > 0:
                    sig_celltypes_perspective1.add(ct)

                dotplot(
                    df, title=f"{title} — {ct} ({direction}) — {gs_label}",
                    outpath=tag_fig_dir / f"dotplot_{ct.replace(' ','_')}_{direction}_{gs_label}",
                )

            sections_by_celltype.setdefault(ct, {})[direction] = res

    # Excel sheet per comparison (one sheet per cell type, GO_BP shown; Reactome in CSV)
    for ct, dirs in sections_by_celltype.items():
        sections = []
        for direction, hdr, fill in [("up", HDR_UP, FILL_UP), ("down", HDR_DOWN, FILL_DOWN)]:
            df = dirs.get(direction, {}).get("GO_BP", pd.DataFrame())
            sections.append((f"{direction.upper()} — GO_BP (top 15)", df, hdr, fill))
        sheet_name = f"{tag[:15]}_{ct[:10]}"
        write_sheet(wb, sheet_name, sections)

out_xlsx = FIG_ENR_DIR / "enrichment_strategic.xlsx"
wb.save(out_xlsx)
print(f"\n✓ Workbook salvo: {out_xlsx}")

# =============================================================================
# SUMMARY
# =============================================================================

summary_df = pd.DataFrame(all_summary_rows)
summary_df.to_csv(ENRICH_DIR / "enrichment_summary_all.csv", index=False)

print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
if len(summary_df):
    print(summary_df.to_string(index=False))
else:
    print("  Nenhum enrichment rodado (verificar Level 2 gene lists).")

print(f"\n  Tipos celulares significativos na Perspectiva I (usados para restringir I-G1/I-G2): "
      f"{sorted(sig_celltypes_perspective1)}")
print(f"\n  Tabelas  : {ENRICH_DIR}/<tag>/")
print(f"  Figuras  : {FIG_ENR_DIR}/<tag>/")
print(f"  Excel    : {out_xlsx}")
print("\n>>> Pipeline TT4 FAPESP concluido (sem etapa de trajetoria).")
